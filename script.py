import requests
import pandas as pd
import time
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
def get_user_instagram_id(access_token, page_id):
    url = f"https://graph.facebook.com/v22.0/{page_id}?fields=instagram_business_account&access_token={access_token}"
    response = requests.get(url)
    if response.status_code == 200:
        instagram_account = response.json()
        return instagram_account["instagram_business_account"]["id"]
    else:
        print(f"Failed to retrieve instagram account: {response.status_code}")
        print(f"Response content: {response.content}")
        return None

def get_page_id(access_token):
    url = f"https://graph.facebook.com/v22.0/me/accounts?fields=id%2Cname%2Caccess_token&access_token={access_token}"
    response = requests.get(url)
    if response.status_code == 200:
        pages = response.json()
        return pages["data"][0]["id"]
    else:
        print(f"Failed to retrieve page id: {response.status_code}")
        print(f"Response content: {response.content}")
        return None

def get_hashtag_id(access_token, hashtag_name, user_instagram_id):
    url = f"https://graph.facebook.com/v22.0/ig_hashtag_search?user_id={user_instagram_id}&q={hashtag_name}&access_token={access_token}"
    response = requests.get(url)
    if response.status_code == 200:
        hashtag = response.json()
        return hashtag["data"][0]["id"]
    else:
        print(f"Failed to retrieve hashtag id: {response.status_code}")
        print(f"Response content: {response.content}")
        return None

def fetch_recent_media_hashtag(access_token, hashtag_id, user_id, start_url):
    video_data = []
    url = start_url or f"https://graph.facebook.com/v22.0/{hashtag_id}/recent_media?user_id={user_id}&fields=permalink&access_token={access_token}"
    api_call_count = 0
    calls_without_results = 0  # Contatore per le chiamate senza risultati

        # Controlla se sono stati trovati nuovi dati
        if data.get("data", []):
            calls_without_results = 0  # Resetta il contatore
            for media in data.get("data", []):
                if "/reel/" in media["permalink"]:
                    video_data.append({
                        "permalink": media["permalink"],
                    })
                    print(f"REELS INSERTED: {len(video_data)}")
        else:
            calls_without_results += 1
            print(f"No new data found in this call. Calls without results: {calls_without_results}")
            if calls_without_results >= 5:
                print("Stopping fetch after 5 calls without results.")
                break  # Interrompi il ciclo

        # Gestione della paginazione (come prima)
        url = data.get('paging', {}).get('next', None)
        if url:
            print(f"Next page URL: {url}")
        else:
            print("No more pages to fetch.")
            break

        rate_limit_remaining = response.headers.get('X-RateLimit-Remaining', None)
        if rate_limit_remaining and int(rate_limit_remaining) < 200:
            print("Close to call limit, pausing for 3600 seconds...")
            time.sleep(3600)

    return video_data


def save_video_urls_to_excel(video_data, file_name):
    if not video_data:
        print("No video data to save.")
        return

    # Controlla se la cartella esiste e crea se non esiste
    folder_path = os.path.dirname(file_name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
        existing_urls = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Video URL"])
        existing_urls = []

    new_video_data = [video for video in video_data if video["permalink"] not in existing_urls]

    if not new_video_data:
        print("No new video data to save.")
        return

    df = pd.DataFrame(new_video_data)
    df.columns = ["Video URL"]

    for index, row in df.iterrows():
        ws.append([row["Video URL"]])

    ws.column_dimensions[get_column_letter(1)].width = 46

    for row in range(ws.max_row - len(new_video_data) + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.hyperlink = cell.value
        cell.style = "Hyperlink"

    wb.save(file_name)
    print(f"Saved {len(new_video_data)} new video URLs in {file_name}")

def read_access_token(file_path):
    with open(file_path, 'r') as file:
        return file.read().strip()

if __name__ == "__main__":
    access_token = read_access_token("/content/drive/MyDrive/Colab Notebooks/Credentials.txt")
    hashtag_name = "endodontics"
    user_instagram_id = get_user_instagram_id(access_token, get_page_id(access_token))
    hashtag_id = get_hashtag_id(access_token, hashtag_name, user_instagram_id)
    if hashtag_id:
        url = "";
        video_data = fetch_recent_media_hashtag(access_token, hashtag_id, user_instagram_id, url)
    save_video_urls_to_excel(video_data, f"/content/drive/MyDrive/Colab Notebooks/{hashtag_name}_urls.xlsx")
